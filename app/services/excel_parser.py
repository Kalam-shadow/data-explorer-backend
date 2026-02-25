"""Excel parsing utilities with robust header detection and schema inference.

This module focuses on durable Excel parsing for production: better error
handling, logging, improved header detection heuristics, hidden column
detection, and conservative schema inference.
"""

import logging
import re
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple

import pandas as pd
from fastapi import HTTPException
from openpyxl import load_workbook

logger = logging.getLogger(__name__)


# -----------------------------
# Utilities
# -----------------------------


def _normalize_col(name: Any) -> str:
    if name is None:
        return ""
    s = str(name).strip()
    s = re.sub(r"\s+", "_", s)
    s = re.sub(r"[^0-9a-zA-Z_]+", "", s)
    s = re.sub(r"_+", "_", s)
    s = s.strip("_")
    s = s.lower()
    if not s:
        return ""
    if s[0].isdigit():
        s = f"c_{s}"
    return s


def _ensure_path(path: Any) -> Path:
    p = Path(path)
    if not p.exists():
        logger.debug("Path does not exist: %s", p)
        raise HTTPException(status_code=400, detail="File not found")
    return p


def _read_single_sheet(path: Any, sheet_name: Optional[str] = None, header: Optional[int] = None) -> pd.DataFrame:
    p = _ensure_path(path)

    try:
        # Validate sheet exists when a name is provided
        if sheet_name is not None:
            try:
                wb = load_workbook(filename=str(p), read_only=True, data_only=True)
                if sheet_name not in wb.sheetnames:
                    wb.close()
                    raise HTTPException(status_code=400, detail=f"Sheet '{sheet_name}' not found")
                wb.close()
            except HTTPException:
                raise
            except Exception as e:  # pragma: no cover - IO/runtime issues
                logger.exception("Error validating workbook: %s", e)
                raise HTTPException(status_code=400, detail=f"Failed to open workbook: {e}")

        df = pd.read_excel(
            str(p),
            sheet_name=sheet_name,
            header=header,
            engine="openpyxl",
        )
    except HTTPException:
        raise
    except Exception as e:  # pragma: no cover - pandas/read errors
        logger.exception("Failed to read Excel file: %s", e)
        raise HTTPException(status_code=400, detail=f"Failed to read Excel file: {e}")

    if isinstance(df, dict):
        if not df:
            raise HTTPException(status_code=400, detail="Excel file contains no sheets")
        df = next(iter(df.values()))

    if df is None or getattr(df, "empty", True):
        raise HTTPException(status_code=400, detail="Excel sheet is empty")

    return df


# -----------------------------
# Smart Header Detection
# -----------------------------


def _detect_header_row(path: Any, sheet_name: Optional[str] = None, max_scan_rows: int = 20) -> int:
    """Detect the most likely header row index (0-based) within the first
    `max_scan_rows` rows.
    """
    df_raw = _read_single_sheet(path, sheet_name=sheet_name, header=None)
    df_raw = df_raw.head(max_scan_rows)
    df_raw = df_raw.dropna(how="all")

    if df_raw.empty:
        raise HTTPException(status_code=400, detail="Excel contains only empty rows")

    best_score = float("-inf")
    best_index = 0

    for idx, row in df_raw.iterrows():
        values = row.dropna()
        if values.empty:
            continue

        # counts
        text_count = sum(isinstance(v, str) for v in values)
        numeric_count = sum(isinstance(v, (int, float)) for v in values)
        nonempty_ratio = len(values) / max(1, len(row))

        # uniqueness among string values (headers tend to be unique)
        str_vals = [str(v).strip() for v in values if isinstance(v, str) and str(v).strip()]
        unique_ratio = (len(set(str_vals)) / len(str_vals)) if str_vals else 0.0

        # Basic heuristic: prefer rows with many strings, high uniqueness and many non-empty cells
        score = (text_count - numeric_count) + (unique_ratio * 2.0) + (nonempty_ratio * 0.5)

        if score > best_score:
            best_score = score
            best_index = int(idx)

    return int(best_index)


# -----------------------------
# Hidden Columns Detection
# -----------------------------


def _hidden_columns(path: Any, sheet_name: Optional[str] = None) -> List[int]:
    p = Path(path)
    try:
        wb = load_workbook(filename=str(p), read_only=True, data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        hidden = []

        def col_letter_to_index(letter: str) -> int:
            idx = 0
            for ch in letter.upper():
                if "A" <= ch <= "Z":
                    idx = idx * 26 + (ord(ch) - ord("A") + 1)
            return idx - 1

        # column_dimensions keys are column letters (e.g., 'A', 'B', 'AA')
        for col_letter, col_dim in ws.column_dimensions.items():
            try:
                if getattr(col_dim, "hidden", False):
                    hidden.append(col_letter_to_index(col_letter))
            except Exception:
                continue

        wb.close()
        return hidden
    except Exception as e:  # pragma: no cover - IO/runtime
        logger.debug("Could not determine hidden columns: %s", e)
        return []


# -----------------------------
# Schema Inference
# -----------------------------


def _infer_schema(df: pd.DataFrame) -> Dict[str, str]:
    schema: Dict[str, str] = {}
    for col in df.columns:
        series = df[col].dropna()
        # sample up to 200 non-null values to infer
        sample = series.iloc[:200]

        if sample.empty:
            schema[col] = "string"
            continue

        if pd.api.types.is_bool_dtype(sample) or sample.astype(str).str.lower().isin(["true", "false", "0", "1"]).all():
            schema[col] = "boolean"
        elif pd.api.types.is_integer_dtype(sample) or pd.api.types.is_float_dtype(sample):
            schema[col] = "number"
        elif pd.api.types.is_datetime64_any_dtype(sample):
            schema[col] = "datetime"
        else:
            # fallback: try parsing datetimes on sample
            try:
                parsed = pd.to_datetime(sample, errors="coerce")
                if parsed.notna().sum() / max(1, len(parsed)) > 0.8:
                    schema[col] = "datetime"
                    continue
            except Exception:
                pass
            schema[col] = "string"

    return schema

def _expand_merged_cells(path, sheet_name=None):
    """Expand merged cells in-place so downstream pandas reads repeated values.

    Fills each merged range with the value from its top-left cell and saves
    the workbook back to the same path. Non-fatal on error (logs and returns).
    """
    p = _ensure_path(path)
    try:
        wb = load_workbook(filename=str(p), data_only=True)
        ws = wb[sheet_name] if sheet_name else wb.active

        # Iterate over a copy of the ranges to avoid mutation issues
        for merged_range in list(ws.merged_cells.ranges):
            min_col, min_row, max_col, max_row = merged_range.bounds
            value = ws.cell(row=min_row, column=min_col).value

            for row in range(min_row, max_row + 1):
                for col in range(min_col, max_col + 1):
                    ws.cell(row=row, column=col).value = value

        wb.save(str(p))
        wb.close()
    except Exception as e:  # pragma: no cover - IO/runtime
        logger.debug("Could not expand merged cells: %s", e)
        return

def _detect_header_block(path, sheet_name=None, max_scan_rows=20):
    """Detect a header block start and depth (number of header rows).

    Returns a tuple (header_start_index, header_depth). If no clear header
    is found returns (0, 1).
    """
    df_raw = _read_single_sheet(path, sheet_name=sheet_name, header=None)
    df_raw = df_raw.head(max_scan_rows)

    # drop completely empty rows to avoid noise
    df_raw = df_raw.dropna(how="all")

    header_start = None
    for idx, row in df_raw.iterrows():
        non_null = [v for v in row if pd.notna(v)]
        if not non_null:
            continue
        text_ratio = sum(isinstance(v, str) for v in non_null) / max(1, len(non_null))
        if text_ratio > 0.6:
            header_start = int(idx)
            break

    if header_start is None:
        return 0, 1

    # Check the next row to see if it's also header-like (multi-row header)
    next_idx = header_start + 1
    if next_idx < len(df_raw):
        next_row = df_raw.iloc[next_idx]
        non_null_next = [v for v in next_row if pd.notna(v)]
        next_ratio = sum(isinstance(v, str) for v in non_null_next) / max(1, len(non_null_next))
        if next_ratio > 0.5:
            return header_start, 2

    return header_start, 1


def _is_subject_triplet_format(df: pd.DataFrame) -> bool:
    """Return True if the first row looks like subject triplet sub-headers (Code/GR/GP).

    Heuristic: the first row contains repeated tokens like 'gr', 'gp', 'code', or 'sub'.
    Requires at least three matches to consider it a triplet layout.
    """
    try:
        first_row = df.iloc[0].astype(str).str.lower()
    except Exception:
        return False

    # normalize punctuation and whitespace for simpler matching
    first_row = first_row.str.replace(r"[\.\#]", "", regex=True).str.strip()
    matches = first_row.str.contains(r"\b(gr|gp|code|sub)\b", regex=True)
    return int(matches.sum()) >= 3



def _parse_subject_triplet(df_raw: pd.DataFrame, header_start: int = 0) -> Tuple[pd.DataFrame, List[List[Optional[str]]]]:
    """Parse a DataFrame in the subject-triplet layout.

    - Uses `header_start` row as sub-header (e.g., 'Sub #1', 'GR.', 'GP.')
    - Combines each column's generated base name with the sub-header value
      to form deterministic column identifiers.

    Returns (parsed_df, original_header_rows)
    """
    # sub-header row
    sub_header = df_raw.iloc[header_start].tolist()
    data = df_raw.iloc[header_start + 1 :].reset_index(drop=True)

    # Generate base column names (use existing df column labels if they are strings,
    # otherwise fallback to positional names)
    base_cols: List[str] = []
    for c in df_raw.columns:
        if isinstance(c, str) and c.strip():
            base_cols.append(c)
        else:
            base_cols.append(f"col_{int(c)}")

    new_cols: List[str] = []
    for base, sub in zip(base_cols, sub_header):
        if pd.isna(sub) or str(sub).strip() == "":
            combined = base
        else:
            combined = f"{base}_{str(sub).strip()}"
        nc = _normalize_col(combined)
        if not nc:
            nc = f"{base}"
        new_cols.append(nc)

    data.columns = new_cols

    # build original header rows for metadata (single-row sub-header)
    original_header_rows: List[List[Optional[str]]] = [[None if pd.isna(v) else str(v).strip() for v in sub_header]]

    # drop fully-empty subject blocks (columns) conservatively
    data = data.dropna(axis=1, how="all")

    return data, original_header_rows


# -----------------------------
# Main Parser
# -----------------------------


def parse_excel(
    path: Any,
    sheet_name: Optional[str] = None,
    required_columns: Optional[List[str]] = None,
    max_scan_rows: int = 20,
) -> Tuple[pd.DataFrame, Dict[str, Any]]:
    """Parse an Excel file and return a cleaned DataFrame plus metadata.

    - Detects header row automatically within the first `max_scan_rows` rows.
    - Normalizes column names to safe identifiers.
    - Removes empty rows/columns and hidden columns.
    - Infers a simple schema for each column.
    """
    p = _ensure_path(path)
    
    # Expand merged cells first so pandas doesn't get NaNs for merged headers
    _expand_merged_cells(p, sheet_name=sheet_name)

    # Quick preview: check for subject-triplet layout (Code/GR/GP per subject)
    try:
        preview = _read_single_sheet(p, sheet_name=sheet_name, header=0)
        if _is_subject_triplet_format(preview):
            # deterministic parsing for subject-triplet layout
            df_parsed, original_header_rows = _parse_subject_triplet(preview, header_start=0)

            # Remove hidden columns if any (index mapping applies to parsed df)
            hidden_idx = _hidden_columns(p, sheet_name)
            if hidden_idx:
                cols_to_drop = [df_parsed.columns[i] for i in hidden_idx if 0 <= i < len(df_parsed.columns)]
                if cols_to_drop:
                    logger.debug("Dropping hidden columns (triplet mode): %s", cols_to_drop)
                    df_parsed = df_parsed.drop(columns=cols_to_drop, errors="ignore")

            # Normalize missing values and trim strings
            df_parsed = df_parsed.where(pd.notnull(df_parsed), None)
            for col in df_parsed.select_dtypes(include=[object]).columns:
                df_parsed[col] = df_parsed[col].apply(lambda v: v.strip() if isinstance(v, str) else v)

            # Required column check
            missing: List[str] = []
            if required_columns:
                norm_required = [_normalize_col(c) for c in required_columns]
                for orig, norm in zip(required_columns, norm_required):
                    if norm not in df_parsed.columns:
                        missing.append(orig)

            schema = _infer_schema(df_parsed)

            meta: Dict[str, Any] = {
                "header_row": 0,
                "header_depth": 1,
                "original_header_rows": original_header_rows,
                "columns": list(df_parsed.columns),
                "row_count": len(df_parsed),
                "schema": schema,
                "missing_required": missing,
            }

            return df_parsed, meta
    except Exception:
        # if preview fails, fall back to normal detection
        pass

    # Detect header block (start row and depth)
    header_start, header_depth = _detect_header_block(p, sheet_name=sheet_name, max_scan_rows=max_scan_rows)

    # Read raw with no header so we can combine header rows ourselves
    df = _read_single_sheet(p, sheet_name=sheet_name, header=None)

    # Drop completely empty rows/columns from the raw sheet
    df = df.dropna(axis=0, how="all")
    df = df.dropna(axis=1, how="all")

    if df.empty:
        raise HTTPException(status_code=400, detail="No usable data found")

    # If we detected a header block, combine header rows column-wise
    if header_start is None:
        header_start = 0
        header_depth = 1

    # ensure header indices are within bounds
    max_row_index = len(df) - 1
    if header_start > max_row_index:
        header_start = 0
        header_depth = 1

    # Slice header rows and the data portion
    header_rows = df.iloc[header_start : header_start + header_depth]
    data = df.iloc[header_start + header_depth :].reset_index(drop=True)

    # Combine headers column-wise into composite names
    combined_headers: List[str] = []
    num_cols = df.shape[1]
    for col_idx in range(num_cols):
        parts: List[str] = []
        for r in range(header_rows.shape[0]):
            try:
                val = header_rows.iloc[r, col_idx]
            except Exception:
                val = None
            if pd.notna(val) and str(val).strip():
                parts.append(str(val).strip())
        combined = "_".join(parts)
        nc = _normalize_col(combined)
        if not nc:
            nc = f"col_{col_idx}"
        combined_headers.append(nc)

    data.columns = combined_headers
    df = data

    # Remove hidden columns if any
    hidden_idx = _hidden_columns(p, sheet_name)
    if hidden_idx:
        cols_to_drop = [df.columns[i] for i in hidden_idx if 0 <= i < len(df.columns)]
        if cols_to_drop:
            logger.debug("Dropping hidden columns: %s", cols_to_drop)
            df = df.drop(columns=cols_to_drop, errors="ignore")

    # Normalize missing values to None (JSON-friendly)
    df = df.where(pd.notnull(df), None)

    # Trim string fields
    for col in df.select_dtypes(include=[object]).columns:
        df[col] = df[col].apply(lambda v: v.strip() if isinstance(v, str) else v)

    # Required column check (original names preserved in required_columns param)
    missing: List[str] = []
    if required_columns:
        norm_required = [_normalize_col(c) for c in required_columns]
        for orig, norm in zip(required_columns, norm_required):
            if norm not in df.columns:
                missing.append(orig)

    schema = _infer_schema(df)

    # capture original header rows for user preview
    original_header_rows: List[List[Optional[str]]] = []
    try:
        for r in range(header_rows.shape[0]):
            row_vals = [None if pd.isna(v) else str(v).strip() for v in header_rows.iloc[r].tolist()]
            original_header_rows.append(row_vals)
    except Exception:
        original_header_rows = []

    meta: Dict[str, Any] = {
        "header_row": header_start,
        "header_depth": header_depth,
        "original_header_rows": original_header_rows,
        "columns": list(df.columns),
        "row_count": len(df),
        "schema": schema,
        "missing_required": missing,
    }

    return df, meta


def _split_tables(df: pd.DataFrame) -> List[pd.DataFrame]:
    """Split a DataFrame into multiple tables separated by fully-empty rows.

    Returns a list of DataFrames (each may need header handling).
    """
    tables: List[pd.DataFrame] = []
    current_rows: List[pd.Series] = []

    for _, row in df.iterrows():
        if row.isna().all():
            if current_rows:
                tables.append(pd.DataFrame(current_rows))
                current_rows = []
        else:
            current_rows.append(row)

    if current_rows:
        tables.append(pd.DataFrame(current_rows))

    return tables


__all__ = [
    "parse_excel",
    "_detect_header_row",
    "_detect_header_block",
    "_expand_merged_cells",
    "_hidden_columns",
    "_split_tables",
]

